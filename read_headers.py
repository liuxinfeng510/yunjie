# -*- coding: utf-8 -*-
import openpyxl
import sys, os
sys.stdout.reconfigure(encoding='utf-8')

path = os.path.join('F:\\', '\u836f\u623f\u7ba1\u7406\u7cfb\u7edf\u5b9e\u65bd', '\u8363\u6cf0\u5927\u836f\u623f', '\u6279\u6b21\u5e93\u5b58\u67e5\u8be2.xlsx')
wb = openpyxl.load_workbook(path, read_only=True)
ws = wb.active
for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
    for cell in row:
        print(f'Col {cell.column - 1}: {cell.value}')
print('---')
for row in ws.iter_rows(min_row=2, max_row=2, values_only=False):
    for cell in row:
        print(f'  Row2 Col {cell.column - 1}: {cell.value}')
wb.close()
